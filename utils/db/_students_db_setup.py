import sqlite3
import re
from openpyxl import load_workbook


if __name__ == '__main__':
    connector = sqlite3.connect('students.sqlite3')

    cursor = connector.cursor()

    with connector:
        cursor.execute("""CREATE TABLE streams(
                          id INTEGER PRIMARY KEY,
                          name TEXT,
                          course INTEGER
                          )""")

        cursor.execute("""CREATE TABLE students(
                          name TEXT,
                          surname TEXT,
                          patronymic TEXT,
                          total REAL,
                          stream_id INTEGER REFERENCES streams(id)
                          )""")


    book = load_workbook('/home/ivan/Downloads/Telegram Desktop/рейтинг.xlsx')
    sheet = book.worksheets[0]


    specialty, course = None, None
    for row in sheet.iter_rows(values_only=True):
        if not isinstance(row[0], int):
            searched_col = None

            for value in row:
                if value is not None:
                    searched_col = value
                    break

            if 'Спеціальність' in searched_col:
                specialty = ' '.join(re.findall(r':.+', searched_col)[0].split(' ')[1:])
            else:
                course_num = int(re.findall(r'\d', searched_col)[0])
                course = course_num + 4 if 'маг' in searched_col else course_num
            with connector:
                cursor.execute("INSERT INTO streams (name, course) VALUES (:name, :course)",
                               {'name': specialty, 'course': course})
        else:
            surname, name, patronymic = row[1].split()
            cursor.execute("SELECT * FROM streams WHERE name=:name AND course=:course",
                           {'name': specialty, 'course': course})
            specialty_id = cursor.fetchall()[0][0]
            with connector:
                cursor.execute("""INSERT INTO students (name, surname, patronymic, total, stream_id)
                               VALUES (:name, :surname, :patronymic, :total, :specialty_id)""",
                               {'name': name, 'surname': surname, 'patronymic': patronymic, 'total': row[-1],
                                'specialty_id': specialty_id})


    cursor.execute("""SELECT stud.surname, stud.name, stud.patronymic, stud.total, spec.name 
                   FROM students stud LEFT JOIN streams spec ON spec.id=stud.stream_id 
                   WHERE stud.total >= 90 AND spec.name LIKE 'Прик%' AND spec.course=3 ORDER BY stud.total DESC""")

    res = cursor.fetchall()
    print(len(res))

    for row in res:
        print(row)

    connector.close()